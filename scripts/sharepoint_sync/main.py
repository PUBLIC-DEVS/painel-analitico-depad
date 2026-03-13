import os
import io
import pandas as pd
from dotenv import load_dotenv
import urllib.parse
import msal
import requests
import openpyxl
import unicodedata
import re
import json
from datetime import datetime, date
from sqlalchemy import create_engine, text
from typing import Dict, List, Any
from scripts.sharepoint_sync.models import (
    RawBaseVigente, RawFiscalizacao, RawRenovacao, RawPagamento, RawRepasse,
    clean_cnpj, clean_numeric, clean_date
)
from pydantic import ValidationError

# Carrega variáveis de ambiente
load_dotenv()

# Configurações SharePoint
SP_TENANT_ID = os.environ.get("SHAREPOINT_TENANT_ID", "common")
SP_CLIENT_ID = os.environ.get("SHAREPOINT_CLIENT_ID")
SP_HOSTNAME = os.environ.get("SHAREPOINT_HOSTNAME", "mdsgov.sharepoint.com")
SP_SITE_PATH = os.environ.get("SHAREPOINT_SITE_PATH", "/sites/EquipeDEPAD")

# Flag para rodar usando arquivos locais
USE_LOCAL_FILES = os.environ.get("USE_LOCAL_FILES", "False").lower() == "true"

# Caminhos remotos
SP_PLANILHA_FISCALIZACAO_PATH = os.environ.get("SHAREPOINT_PLANILHA_FISCALIZACAO", "General/DEPAD/Contratos/Fiscalização/Controle de Fiscalização.xlsx")
SP_PLANILHA_BASE_PATH = os.environ.get("SHAREPOINT_PLANILHA_BASE", "General/DEPAD/BASE.xlsx")
SP_PLANILHA_RENOVACAO_PATH = os.environ.get("SHAREPOINT_PLANILHA_RENOV", "General/DEPAD/Contratos/Renovação e contratação/PLANILHA DEPAD FINAL EDITAL.xlsx")
SP_PLANILHA_PAGAMENTOS_PATH = os.environ.get("SHAREPOINT_PLANILHA_PAGAMENTOS", "General/DEPAD/Contratos/Pagamentos e prestação de contas/Planilha Pagamentos.xlsx")
SP_PLANILHA_REPASSES_PATH = os.environ.get("SHAREPOINT_PLANILHA_REPASSES", "General/DEPAD/Repasses/Execução/PLANILHA GERAL DE ACOMPANHAMENTO.xlsx")

# Caminhos locais
LOCAL_PLANILHA_FISCALIZACAO_PATH = os.environ.get("LOCAL_PLANILHA_FISCALIZACAO_PATH", "dados_amostra/Controle de Fiscalização.xlsx")
LOCAL_PLANILHA_BASE_PATH = os.environ.get("LOCAL_PLANILHA_BASE_PATH", "dados_amostra/BASE.xlsx")
LOCAL_PLANILHA_RENOVACAO_PATH = os.environ.get("LOCAL_PLANILHA_RENOVACAO_PATH", "dados_amostra/PLANILHA DEPAD FINAL EDITAL.xlsx")
LOCAL_PLANILHA_PAGAMENTOS_PATH = os.environ.get("LOCAL_PLANILHA_PAGAMENTOS_PATH", "dados_amostra/Planilha Pagamentos.xlsx")
LOCAL_PLANILHA_REPASSES_PATH = os.environ.get("LOCAL_PLANILHA_REPASSES_PATH", "dados_amostra/PLANILHA GERAL DE ACOMPANHAMENTO.xlsx")

# Configurações Neon Database
DATABASE_URL = os.environ.get("DATABASE_URL")

def obter_token_msal():
    """Autentica via MSAL usando fluxo interativo/device e faz cache do token"""
    print("Iniciando processo de autenticação (MSAL)...")
    if not SP_CLIENT_ID:
        raise ValueError("A variável SHAREPOINT_CLIENT_ID está ausente no .env")
        
    authority = f"https://login.microsoftonline.com/{SP_TENANT_ID}"
    scopes = ["Sites.Read.All", "Files.Read.All"]
    cache_arquivo = "token_cache.bin"
    cache = msal.SerializableTokenCache()

    if os.path.exists(cache_arquivo):
        try:
            cache.deserialize(open(cache_arquivo, "r").read())
            print("Cache de autenticação carregado.")
        except Exception as e:
            print(f"Cache corrompido, será recriado. Erro: {e}")

    app = msal.PublicClientApplication(client_id=SP_CLIENT_ID, authority=authority, token_cache=cache)
    contas = app.get_accounts()
    resultado = None

    if contas:
        print("Conta encontrada no cache, obtendo token silenciosamente...")
        resultado = app.acquire_token_silent(scopes, account=contas[0])
    
    if not resultado:
        print("Sua primeira vez ou token expirado. Abrindo navegador para login interativo...")
        resultado = app.acquire_token_interactive(scopes=scopes, timeout=120)

    if resultado and "access_token" in resultado:
        with open(cache_arquivo, "w") as f:
            f.write(cache.serialize())
        print("Autenticação concluída e token cacheado!")
        return resultado["access_token"]

    raise RuntimeError(f"Falha ao obter token: {resultado.get('error_description', str(resultado))}")

def normalize_column_name(col_name: str) -> str:
    """Normaliza o nome de uma coluna removendo acentos, caracteres especiais e padronizando espaços."""
    if not isinstance(col_name, str): 
        col_name = str(col_name)
    s = col_name.strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^A-Z0-9_]+', '_', s)
    s = re.sub(r'_+', '_', s)
    s = s.strip('_')
    return s.lower()

def normalize_str(s: Any) -> str:
    """Normaliza strings para comparação (remover acentos, espaços extras e caixa baixa)."""
    if pd.isna(s) or s is None:
        return ""
    s = str(s).strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def preprocess_df(df: pd.DataFrame, cnpj_source_col: str, dropna_cnpj: bool = False, keep_source_col: bool = False, cnpj_target_field: str = 'cnpj') -> pd.DataFrame:
    df_copy = df.copy()

    # Normalização de todas as colunas para facilitar busca e garantir compatibilidade DB
    df_copy.columns = [normalize_column_name(col) for col in df_copy.columns]
    
    # Remove colunas com nomes vazios (comum em planilhas com formatação suja)
    cols_to_drop = [c for c in df_copy.columns if not c or c.strip() == ""]
    if cols_to_drop:
        print(f"Limpando {len(cols_to_drop)} colunas sem nome identificadas.")
        df_copy.drop(columns=cols_to_drop, inplace=True)

    cnpj_target_field_norm = normalize_column_name(cnpj_target_field)
    cnpj_source_norm = normalize_column_name(cnpj_source_col)
    
    # Identificação e limpeza do CNPJ
    actual_col = None
    if cnpj_target_field_norm in df_copy.columns:
        actual_col = cnpj_target_field_norm
    elif cnpj_source_norm in df_copy.columns:
        actual_col = cnpj_source_norm
    
    if actual_col:
        df_copy[cnpj_target_field] = df_copy[actual_col].ffill().apply(clean_cnpj)
        if actual_col != cnpj_target_field and not keep_source_col:
            df_copy.drop(columns=[actual_col], inplace=True, errors='ignore')
        
        if dropna_cnpj:
            df_copy.dropna(subset=[cnpj_target_field], inplace=True)
            df_copy = df_copy[df_copy[cnpj_target_field].astype(str).str.match(r'^\d{14}$', na=False)]
    else:
        df_copy[cnpj_target_field] = None

    return df_copy

def unmerge_and_fill_excel(excel_content: io.BytesIO, sheet_name: Any) -> io.BytesIO:
    """Desmescla células em um arquivo Excel e preenche os valores mesclados."""
    workbook = openpyxl.load_workbook(excel_content, data_only=False)

    if isinstance(sheet_name, str):
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não encontrada no workbook.")
        sheet = workbook[sheet_name]
    elif isinstance(sheet_name, int):
        sheet = workbook.worksheets[sheet_name]
    else:
        raise TypeError("sheet_name deve ser string ou int.")

    merged_cells_ranges = list(sheet.merged_cells)

    for merged_cell_range in merged_cells_ranges:
        min_col, min_row, max_col, max_row = merged_cell_range.min_col, merged_cell_range.min_row, \
                                             merged_cell_range.max_col, merged_cell_range.max_row
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(merged_cell_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = top_left_cell_value

    output_excel_content = io.BytesIO()
    workbook.save(output_excel_content)
    output_excel_content.seek(0)
    return output_excel_content

def read_excel_source(token: str, item_path: str, local_path: str, header=0, sheet_name=0):
    """Lê do cache local ou baixa via Microsoft Graph API"""
    try:
        if USE_LOCAL_FILES:
            with open(local_path, 'rb') as f:
                content = f.read()
            processed_excel_content = unmerge_and_fill_excel(io.BytesIO(content), sheet_name)
            df = pd.read_excel(processed_excel_content, header=header, sheet_name=sheet_name)
            return df
            
        headers = {'Authorization': 'Bearer ' + token}
        site_identifier = f"{SP_HOSTNAME}:{SP_SITE_PATH}"
        site_url = f"https://graph.microsoft.com/v1.0/sites/{urllib.parse.quote(site_identifier)}"
        site_response = requests.get(site_url, headers=headers, timeout=30)
        site_response.raise_for_status()
        site_id = site_response.json().get('id')
        
        graph_url_download = (f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/"
                              f"{urllib.parse.quote(item_path)}:/content")
        response = requests.get(graph_url_download, headers=headers, timeout=60)
        response.raise_for_status()
        
        processed_excel_content = unmerge_and_fill_excel(io.BytesIO(response.content), sheet_name)
        df = pd.read_excel(processed_excel_content, header=header, sheet_name=sheet_name)
        return df
    except Exception as e:
        print(f"Erro ao ler arquivo ({local_path if USE_LOCAL_FILES else item_path}): {e}")
        raise e

def sync_mirror_mapping(token: str, db_url: str):
    """Executa a sincronização no modo Mirror Mapping: cada planilha vira uma tabela."""
    if not db_url:
        print("DATABASE_URL não configurada.")
        return

    engine_url = db_url.replace("postgres://", "postgresql://", 1) if db_url.startswith("postgres://") else db_url
    engine = create_engine(engine_url)
    
    configs = [
        {
            "name": "raw_base_vigente",
            "sp_path": SP_PLANILHA_BASE_PATH,
            "loc_path": LOCAL_PLANILHA_BASE_PATH,
            "sheet": 'BASE VIGENTE 2024',
            "header": 0,
            "model": RawBaseVigente
        },
        {
            "name": "raw_fiscalizacao",
            "sp_path": SP_PLANILHA_FISCALIZACAO_PATH,
            "loc_path": LOCAL_PLANILHA_FISCALIZACAO_PATH,
            "sheet": 'Base - Fiscalização',
            "header": 2,
            "model": RawFiscalizacao
        },
        {
            "name": "raw_renovacao",
            "sp_path": SP_PLANILHA_RENOVACAO_PATH,
            "loc_path": LOCAL_PLANILHA_RENOVACAO_PATH,
            "sheet": 'ESTADOS GERAL VIGENTES.',
            "header": 1,
            "model": RawRenovacao
        },
        {
            "name": "raw_pagamento",
            "sp_path": SP_PLANILHA_PAGAMENTOS_PATH,
            "loc_path": LOCAL_PLANILHA_PAGAMENTOS_PATH,
            "sheet": 'Dezembro 2022',
            "header": 11,
            "model": RawPagamento
        },
        {
            "name": "raw_repasse",
            "sp_path": SP_PLANILHA_REPASSES_PATH,
            "loc_path": LOCAL_PLANILHA_REPASSES_PATH,
            "sheet": 'TF, Colaboração e Convênios',
            "header": 0,
            "model": RawRepasse
        }
    ]
    
    for cfg in configs:
        print(f"\n>>> Processando: {cfg['name']} ({cfg['sheet']})")
        try:
            df = read_excel_source(token, cfg["sp_path"], cfg["loc_path"], header=cfg["header"], sheet_name=cfg["sheet"])
            
            # Normaliza e limpa o DataFrame
            found_cnpj = False
            for col in df.columns:
                if normalize_str(col) in ['CNPJ', 'CNPJENTIDADE']:
                    df = df.rename(columns={col: 'cnpj'})
                    found_cnpj = True
                    break
            
            if not found_cnpj and cfg['name'] == 'raw_repasse':
                for col in df.columns:
                    if normalize_str(col) == 'CONVENENTE':
                        df = df.rename(columns={col: 'cnpj'})
                        found_cnpj = True
                        break

            # Pré-processamento Mirror: Normaliza colunas e ffill CNPJ
            df_proc = preprocess_df(df, 'cnpj' if found_cnpj else (df.columns[0] if not df.empty else 'none'), 
                                   dropna_cnpj=False)
            
            # Validação e Limpeza via Pydantic (Preservando todas as colunas)
            cleaned_records = []
            for idx, row in df_proc.iterrows():
                row_dict = row.to_dict()
                # Remove valores nulos do dict para não sobrecarregar
                row_dict = {k: v for k, v in row_dict.items() if pd.notnull(v)}
                # Limpa as colunas conhecidas e mantém as outras
                cleaned_record = cfg["model"].clean_record(row_dict)
                cleaned_records.append(cleaned_record)
            
            if not cleaned_records:
                print(f"ERRO: Nenhum registro processado para {cfg['name']}.")
                continue
                
            df_final = pd.DataFrame(cleaned_records)
            
            print(f"Salvando {len(df_final)} registros na tabela {cfg['name']}...")
            with engine.begin() as conn:
                # Usa replace para criar/recriar a tabela com as colunas do modelo + additional_data
                df_final.to_sql(cfg['name'], conn, if_exists='replace', index=False)
            print(f"Sucesso: {cfg['name']} sincronizada.")
            
        except Exception as e:
            print(f"ERRO ao processar {cfg['name']}: {e}")

def main():
    print(f"Iniciando Sincronização SharePoint -> NeonDB (Modo Mirror Mapping)...")
    token = None
    if not USE_LOCAL_FILES:
        try:
            token = obter_token_msal()
        except Exception as e:
            print(f"Erro de autenticação: {e}")
            return

    try:
        sync_mirror_mapping(token, DATABASE_URL)
        print("\nSincronização concluída com sucesso!")
    except Exception as e:
        print(f"\nERRO CRÍTICO no processo: {e}")

if __name__ == "__main__":
    main()
