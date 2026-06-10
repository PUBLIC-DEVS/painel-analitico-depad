#!/usr/bin/env python3
"""Monta a base UNIFICADA (a "master"): cruza as contratadas (lib/data/base.json)
com os repasses/termos de fomento (planilha de repasses) deduplicando por CNPJ.

Cada comunidade vira UMA linha marcada como Contratada, Repasse ou Ambos — é o que
permite o split "contratada vs geral". Saídas:
  - lib/data/comunidades.json  (consumido pela Base de dados do app)
  - <xlsx de saída>            (a planilha master pra subir no SharePoint)

Uso: python build-master.py <repasses.xlsx> <saida_master.xlsx>
"""
import openpyxl, sys, json, re, os
from collections import defaultdict

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def so_digitos(cnpj): return re.sub(r"\D", "", cnpj or "")
def txt(v): return "" if v is None else str(v).strip()
def to_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if "," in s: s = s.replace(".", "").replace(",", ".")
    try: return float(s)
    except: return 0.0
def brl(v):
    return "R$ " + f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# ── contratadas (já convertidas) ──
with open(os.path.join(REPO, "lib", "data", "base.json"), encoding="utf-8") as f:
    contratadas = json.load(f)

# ── repasses (planilha) ──
repasses_path = sys.argv[1]
wb = openpyxl.load_workbook(repasses_path, read_only=True, data_only=True)
ws = wb.worksheets[0]
rep_rows = list(ws.iter_rows(values_only=True))
wb.close()
# cabeçalho: Nome Parlamentar | Modalidade | Valor de Repasse | Ano | Nome Proponente | CNPJ
repasses = []
for r in rep_rows[1:]:
    if not r or not txt(r[5] if len(r) > 5 else ""): continue
    repasses.append({
        "parlamentar": txt(r[0]), "modalidade": txt(r[1]), "valor": to_float(r[2]),
        "ano": txt(r[3]), "proponente": txt(r[4]), "cnpj": txt(r[5]),
    })

# ── merge por CNPJ ──
master = {}  # chave = cnpj normalizado (ou id único se sem cnpj)
sem_cnpj = 0

def ensure(key):
    if key not in master:
        master[key] = {
            "cnpj": "", "nome": "", "uf": "", "cidade": "",
            "tem_contrato": False, "tem_repasse": False,
            "vagas_contratadas": 0, "adulto_masc": 0, "adulto_feminino": 0, "maes": 0,
            "recurso_anual_contrato": 0.0, "status_ct": "", "ano_contrato": "",
            "valor_repasse_total": 0.0, "qtd_emendas": 0,
            "parlamentares": set(), "modalidades": set(),
        }
    return master[key]

for c in contratadas:
    key = so_digitos(c["cnpj"]) or f"sc-{sem_cnpj}"
    if not so_digitos(c["cnpj"]): sem_cnpj += 1
    m = ensure(key)
    m["tem_contrato"] = True
    m["cnpj"] = c["cnpj"] or m["cnpj"]
    m["nome"] = m["nome"] or c["razao_social"] or c["nome_fantasia"]
    m["uf"] = m["uf"] or c["uf"]
    m["cidade"] = m["cidade"] or c["cidade"]
    m["vagas_contratadas"] += c["vagas_contratadas"] or 0
    m["adulto_masc"] += c["adulto_masc"] or 0
    m["adulto_feminino"] += c["adulto_feminino"] or 0
    m["maes"] += c["maes"] or 0
    m["recurso_anual_contrato"] += to_float(c["recurso_anual"])
    m["status_ct"] = m["status_ct"] or c["status_ct"]
    m["ano_contrato"] = m["ano_contrato"] or str(c["ano"] or "")

for r in repasses:
    key = so_digitos(r["cnpj"]) or f"sr-{sem_cnpj}"
    if not so_digitos(r["cnpj"]): sem_cnpj += 1
    m = ensure(key)
    m["tem_repasse"] = True
    m["cnpj"] = m["cnpj"] or r["cnpj"]
    m["nome"] = m["nome"] or r["proponente"]
    m["valor_repasse_total"] += r["valor"]
    m["qtd_emendas"] += 1
    if r["parlamentar"]: m["parlamentares"].add(r["parlamentar"])
    if r["modalidade"]: m["modalidades"].add(r["modalidade"])

# ── finaliza ──
def tipo(m):
    if m["tem_contrato"] and m["tem_repasse"]: return "Ambos"
    return "Contratada" if m["tem_contrato"] else "Repasse"

unificado = []
for m in master.values():
    unificado.append({
        "cnpj": m["cnpj"], "nome": m["nome"], "uf": m["uf"], "cidade": m["cidade"],
        "tipo": tipo(m), "tem_contrato": m["tem_contrato"], "tem_repasse": m["tem_repasse"],
        "vagas_contratadas": m["vagas_contratadas"], "adulto_masc": m["adulto_masc"],
        "adulto_feminino": m["adulto_feminino"], "maes": m["maes"],
        "recurso_anual_contrato": round(m["recurso_anual_contrato"], 2),
        "valor_repasse_total": round(m["valor_repasse_total"], 2),
        "qtd_emendas": m["qtd_emendas"], "status_ct": m["status_ct"], "ano_contrato": m["ano_contrato"],
        "parlamentares": "; ".join(sorted(m["parlamentares"])),
        "modalidades": "; ".join(sorted(m["modalidades"])),
    })
unificado.sort(key=lambda x: (x["uf"], x["nome"]))

with open(os.path.join(REPO, "lib", "data", "comunidades.json"), "w", encoding="utf-8") as f:
    json.dump(unificado, f, ensure_ascii=False)

# ── master .xlsx ──
out = openpyxl.Workbook()
sh = out.active
sh.title = "Base Unificada"
cols = ["CNPJ", "Nome", "UF", "Cidade", "Tipo", "Tem Contrato", "Tem Repasse",
        "Vagas Contratadas", "Adulto Masc", "Adulto Fem", "Mães",
        "Recurso Anual (Contrato)", "Valor Repasse Total", "Qtd Emendas",
        "Status CT", "Ano", "Parlamentares", "Modalidades"]
sh.append(cols)
for u in unificado:
    sh.append([
        u["cnpj"], u["nome"], u["uf"], u["cidade"], u["tipo"],
        "Sim" if u["tem_contrato"] else "Não", "Sim" if u["tem_repasse"] else "Não",
        u["vagas_contratadas"], u["adulto_masc"], u["adulto_feminino"], u["maes"],
        brl(u["recurso_anual_contrato"]), brl(u["valor_repasse_total"]), u["qtd_emendas"],
        u["status_ct"], u["ano_contrato"], u["parlamentares"], u["modalidades"],
    ])
out.save(sys.argv[2])

# ── relatório ──
n = len(unificado)
amb = sum(1 for u in unificado if u["tipo"] == "Ambos")
ctr = sum(1 for u in unificado if u["tipo"] == "Contratada")
rep = sum(1 for u in unificado if u["tipo"] == "Repasse")
print(f"comunidades únicas (por CNPJ): {n}")
print(f"  só contratadas: {ctr} | só repasse: {rep} | AMBOS: {amb}")
print(f"contratadas na entrada: {len(contratadas)} | linhas de repasse: {len(repasses)}")
print(f"vagas contratadas (total): {sum(u['vagas_contratadas'] for u in unificado):,}")
print(f"repasse total: {brl(sum(u['valor_repasse_total'] for u in unificado))}")
print(f"saídas: lib/data/comunidades.json + {sys.argv[2]}")
